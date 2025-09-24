"""
check_labels.py

Checks interview spreadsheet headers for:
1) Labels not in the canonical Main Question List.
2) Broken Q/R alternation (two consecutive Q_ or two consecutive R_).
3) Embedded artifacts (e.g., jammed headers separated by a TAB).
4) Q/R PAIR MISMATCH: for the same index, Q_i_Label must match R_i_Label.

Marks problems RED and writes a JSON report. If --baseline is provided,
prints the % of errors fixed since that baseline.

Usage (with your paths):
  python check_labels.py \
    --input "/Users/buzhaoliu/Developer/NLP/interview_data_labeled.xlsx" \
    --output "/Users/buzhaoliu/Developer/NLP/interview_data_checked.xlsx" \
    --report "/Users/buzhaoliu/Developer/NLP/check_report.json" \
    [--baseline "/Users/buzhaoliu/Developer/NLP/previous_report.json"]
"""

import os
import re
import json
import argparse
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---- Canonical labels (from Main_Questions_Labels.docx) ----
MAIN_LABELS = {
    "Intro","Current_Status","Wife_Or_You","End_Year","Daily_Tasks",
    "Ward_Languages","Mother_Tongue","Other_Lang_Spoken","Other_Lang_Understood",
    "Hindi_Dialects","Hindi_Dialect_Differences","Correct_Hindi",
    "Dialect_Job_Impact","Dialect_Constituents_1on1_Impact","Dialect_Constituents_Groups_Impact",
    "Dialect_Corp_Impact","Dialect_Leaders_Impact","Dialect_Choice_Example",
    "Dialect_Choice_Grievance","Dialect_Choice_Campaigning","Dialect_Choice_Meetings",
    "Dialect_Choice_Bureaucrats","Dialect_Choice_Other","Dialect_Choice_Multiple",
    "Unmatched","Thanks",
}

RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_RE = re.compile(r"^\s*([RQ])\s*_(\d+)\s*_(.+?)\s*$")

def parse_header(cell_value):
    """Parse 'Q_12_Label' / 'R_7_Label' -> (kind, idx, normalized_label, parse_error)."""
    if cell_value is None:
        return None, None, None, False
    text = str(cell_value).strip()
    parse_error = False

    # Jammed header detection (tabs often mean two headers got merged in one cell)
    if ("\t" in text and "R_" in text) or ("\t" in text and "Q_" in text):
        parse_error = True
        text = text.split("\t")[0].strip()

    m = HEADER_RE.match(text)
    if not m:
        return None, None, None, parse_error

    kind, idx, label = m.group(1), m.group(2), m.group(3)
    # Normalize label whitespace and NBSPs to avoid false negatives
    label = label.strip().replace("\u00A0", " ")
    return kind, idx, label, parse_error

def check_sheet(input_path, output_path):
    wb = load_workbook(input_path)
    ws = wb.active
    max_row, max_col = ws.max_row, ws.max_column

    errors = defaultdict(int)
    total_headers = 0

    # Iterate over header rows only (1-based: rows 1,3,5,...)
    for r in range(1, max_row + 1, 2):
        prev_kind = None
        q_pos, r_pos = {}, {}  # idx -> (col, label)

        # Skip first 3 metadata columns (Name, Date, Location)
        for c in range(4, max_col + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if not val or str(val).strip() == "":
                continue

            kind, idx, label, parse_error = parse_header(val)
            if kind is None:
                continue  # not a Q/R header cell

            total_headers += 1

            # (1) Alternation: should go Q,R,Q,R...
            if prev_kind == kind:
                cell.fill = RED
                errors["alternation"] += 1
                # also mark the previous cell to make the pair obvious
                prev_cell = ws.cell(row=r, column=c-1)
                if prev_cell.value:
                    prev_cell.fill = RED
            prev_kind = kind

            # (2) Embedded artifacts
            if parse_error:
                cell.fill = RED
                errors["embedded"] += 1

            # (3) Label validity
            if label not in MAIN_LABELS:
                cell.fill = RED
                errors["unknown_label"] += 1

            # (4) Collect for Q/R label mismatch check
            if kind == "Q":
                q_pos[idx] = (c, label)
            else:  # 'R'
                r_pos[idx] = (c, label)

        # (4) Q/R pair mismatch: for same index, labels must match
        for idx in set(q_pos) & set(r_pos):
            qc, ql = q_pos[idx]
            rc, rl = r_pos[idx]
            if ql != rl:
                ws.cell(row=r, column=qc).fill = RED
                ws.cell(row=r, column=rc).fill = RED
                errors["qr_mismatch"] += 1

    wb.save(output_path)

    report = {
        "input": input_path,
        "output": output_path,
        "total_header_cells_examined": total_headers,
        "error_counts": dict(errors),
        "total_errors": int(sum(errors.values())),
    }
    return report

def compute_improvement(current_report, baseline_report):
    """Return overall % reduction and per-type reductions since baseline."""
    improvements = {}
    base_total = baseline_report.get("total_errors", 0)
    curr_total = current_report.get("total_errors", 0)

    overall = None
    if base_total > 0:
        overall = round((base_total - curr_total) * 100.0 / base_total, 2)

    base_types = baseline_report.get("error_counts", {})
    curr_types = current_report.get("error_counts", {})
    per_type = {}
    for k in set(base_types) | set(curr_types):
        b = base_types.get(k, 0)
        c = curr_types.get(k, 0)
        per_type[k] = (round((b - c) * 100.0 / b, 2) if b > 0 else None)

    improvements["overall_percent_fixed"] = overall
    improvements["per_type_percent_fixed"] = per_type
    return improvements

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",
        default="/Users/buzhaoliu/Developer/NLP/interview_data_labeled.xlsx",
        help="Path to the input Excel file.")
    parser.add_argument("--output",
        default=None,
        help="Path for the checked Excel file. If omitted, '*_checked.xlsx' is used next to the input.")
    parser.add_argument("--report",
        default=None,
        help="Path to write JSON report. If omitted, 'check_report.json' is written next to the input.")
    parser.add_argument("--baseline",
        default=None,
        help="Optional JSON report from a previous run to compare improvements.")
    args = parser.parse_args()

    # Derive default output/report paths if not provided
    in_base, in_ext = os.path.splitext(args.input)
    output_path = args.output or (in_base + "_checked" + in_ext)
    report_path = args.report or (os.path.join(os.path.dirname(args.input), "check_report.json"))

    current_report = check_sheet(args.input, output_path)

    if args.baseline:
        try:
            with open(args.baseline, "r") as f:
                baseline = json.load(f)
            improvement = compute_improvement(current_report, baseline)
            current_report["comparison_to_baseline"] = improvement
            overall = improvement.get("overall_percent_fixed")
            if overall is not None:
                print(f"✅ Improvement since baseline: {overall}% of errors fixed.")
            else:
                print("ℹ️ Baseline had zero errors or was missing totals.")
        except Exception as e:
            print("⚠️ Could not read/compare baseline report:", e)

    with open(report_path, "w") as f:
        json.dump(current_report, f, indent=2)

    print("✅ Saved checked file:", current_report["output"])
    print("📝 Saved report:", report_path)
    print("Summary:", current_report["error_counts"], "| Total:", current_report["total_errors"])

if __name__ == "__main__":
    main()
