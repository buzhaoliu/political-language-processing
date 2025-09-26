import pandas as pd
from openpyxl import load_workbook

input_file = "/Users/buzhaoliu/Developer/NLP/interview_data_merged.xlsx"
output_file = "/Users/buzhaoliu/Developer/NLP/interview_data_formatted.xlsx"

# Load your Excel (with headers=False if yours is 2-row blocks)
df = pd.read_excel(input_file)

# Try to parse any column named "Date" (case-insensitive)
for col in df.columns:
    if str(col).strip().lower() == "date":
        df[col] = pd.to_datetime(df[col], errors="coerce")

# Save first
df.to_excel(output_file, index=False)

# Now reopen with openpyxl to set the date display format
wb = load_workbook(output_file)
ws = wb.active

for col in range(1, ws.max_column + 1):
    header = str(ws.cell(row=1, column=col).value or "").strip().lower()
    if header == "date":
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col)
            if cell.value:
                cell.number_format = "d mmm"  # e.g. 24 Aug

wb.save(output_file)
print(f"✅ Saved with formatted dates at {output_file}")
